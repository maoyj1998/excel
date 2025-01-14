import argparse
from openpyxl import load_workbook


parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('--in_xlsx', type=str, help='input xlsx file')
parser.add_argument('--out_txt', type=str, help='output txt file')
args = parser.parse_args()


wb = load_workbook(args.in_xlsx)

# 选择工作表
with open(args.out_txt, "w") as fp:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    
        # 遍历所有单元格
        for row in ws.iter_rows():
            for cell in row:
                # 检查单元格是否有批注
                if cell.comment:
                    # 打印批注文本
                    s = f"Sheet: {sheet_name}, Cell: {cell.coordinate}, Comment: {cell.comment.text if cell.comment else 'No comment'}"
                    print(s)
                    fp.write(s)
                    fp.write("\n")
                    fp.write("\n")